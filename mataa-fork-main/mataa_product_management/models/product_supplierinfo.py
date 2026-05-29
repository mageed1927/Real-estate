from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from markupsafe import Markup


class ProductSupplierInfo(models.Model):
    _inherit = ['product.supplierinfo', 'mail.thread']
    _name = 'product.supplierinfo'

    price = fields.Float(tracking=True)
    min_qty = fields.Float(tracking=True)

    def write(self, vals):
        tracked_fields = {'price', 'min_qty'}
        field_labels = {
            'price': _('Vendor Price'),
            'min_qty': _('Vendor Quantity')
        }

        old_values = {
            rec.id: {'price': rec.price, 'min_qty': rec.min_qty}
            for rec in self
        }

        res = super(ProductSupplierInfo, self).write(vals)

        for record in self:
            changes = []
            for field in tracked_fields:
                old_val = old_values[record.id][field]
                new_val = record[field]
                if old_val != new_val:
                    label = field_labels[field]
                    changes.append(Markup("<li><strong>%(label)s:</strong> %(old)s → %(new)s</li>") % {
                        'label': label,
                        'old': old_val,
                        'new': new_val
                    })

            if changes:
                partner_link = record.partner_id._get_html_link()
                base_message = Markup(
                    "<div style='margin-bottom: 8px;'>"
                    "<p style='margin-bottom: 4px;'>%(partner_link)s updated:</p>"
                    "<ul style='margin-left: 16px;'>%(changes)s</ul>"
                    "</div>"
                ) % {
                                   'partner_link': partner_link,
                                   'changes': Markup(''.join(changes))
                               }

                template_message = base_message
                if record.product_id:
                    variant_link = record.product_id._get_html_link()
                    template_message = Markup(
                        "<div style='margin-bottom: 8px;'>"
                        "<p style='margin-bottom: 4px;'>%(partner_link)s updated for %(variant_link)s:</p>"
                        "<ul style='margin-left: 16px;'>%(changes)s</ul>"
                        "</div>"
                    ) % {
                                           'partner_link': partner_link,
                                           'variant_link': variant_link,
                                           'changes': Markup(''.join(changes))
                                       }

                    record.product_id.message_post(
                        body=base_message,
                        message_type='comment',
                        subtype_xmlid='mail.mt_note'
                    )

                record.product_tmpl_id.message_post(
                    body=template_message,
                    message_type='comment',
                    subtype_xmlid='mail.mt_note'
                )
                if record.partner_id.in_house_vendor and record.product_id.get_free_qty() > 0 and record.price != record.product_id.standard_price:
                    # if self._context.get('import_file'):
                    difference_amount = (record.price - record.product_id.standard_price) * record.product_id.get_free_qty()
                    record.product_id.with_context(skip_svl_creation=True).write({"standard_price": record.price})
                    self.env['vendor.credit.difference.tracker'].track_difference(record.partner_id.id, difference_amount)

        return res

    @api.model_create_multi
    def create(self, vals_list):
        records = super(ProductSupplierInfo, self).create(vals_list)
        for record in records:
            if record.partner_id.in_house_vendor and record.product_id.get_free_qty() > 0 and record.price != record.product_id.standard_price:
                # if self._context.get('import_file'):
                difference_amount = (record.price - record.product_id.standard_price) * record.product_id.get_free_qty()
                record.product_id.with_context(skip_svl_creation=True).write({"standard_price": record.price})
                self.env['vendor.credit.difference.tracker'].track_difference(record.partner_id.id, difference_amount)

        return records


    @api.onchange('product_code')
    def _onchange_product_code(self):
        if self.product_code:

            product_template = self.env['product.template'].search([('barcode', '=', self.product_code)], limit=1)
            product_variant = self.env['product.product'].search([('barcode', '=', self.product_code)])

            if product_template:
                self.product_tmpl_id = product_template
                self.product_id = product_variant if product_variant else product_template.product_variant_ids[0]
            else:
                self.product_tmpl_id = False
                self.product_id = False
        else:
            self.product_tmpl_id = False
            self.product_id = False

    @api.constrains('partner_id', 'product_id')
    def _check_duplicate_vendor(self):
        for record in self:
            if not record.partner_id or not record.product_id:
                continue

            domain = [
                ('id', '!=', record.id),
                ('partner_id', '=', record.partner_id.id),
                ('product_id', '=', record.product_id.id)
            ]

            duplicate = self.search(domain, limit=1)
            if duplicate:
                raise ValidationError(_(
                    'Duplicate vendor entry detected!\n'
                    'Vendor "%s" is already linked to the product variant "%s". '
                    'You cannot add the same vendor multiple times for the same product variant.'
                ) % (record.partner_id.display_name, record.product_id.display_name))

    @api.constrains('price', 'product_id', 'product_tmpl_id')
    def _check_purchase_price(self):
        if self.env.context.get('skip_vendor_price_check', False):
            return
        prices_control = self.env['ir.config_parameter'].sudo().get_param('mataa_order_management.prices_control')
        for record in self:
            product = record.product_id or record.product_tmpl_id
            if not product or not record.price:
                continue
            vendor_price = self._get_converted_price(record)
            if bool(prices_control) and vendor_price > min(product.list_price, product.regular_price):
                self._raise_invalid_price_error(record, product, vendor_price)

    def _get_converted_price(self, record):
        product = record.product_id or record.product_tmpl_id
        if not record.currency_id or record.currency_id == product.currency_id:
            return record.price

        return record.currency_id._convert(
            record.price,
            product.currency_id,
            record.company_id or self.env.company,
            fields.Date.today()
        )

    def _raise_invalid_price_error(self, record, product, vendor_price):
        raise ValidationError(_(
            'Invalid purchase price detected!\n'
            'Purchase price (%(purchase_price).2f) cannot be greater than:\n'
            '- Product\'s sale price (%(sale_price).2f)\n'
            '- Product\'s regular price (%(regular_price).2f)\n'
            'Product: %(product)s\n'
            'Vendor: %(vendor)s'
        ) % {
                                  'purchase_price': vendor_price,
                                  'sale_price': product.list_price,
                                  'regular_price': product.regular_price,
                                  'product': product.display_name,
                                  'vendor': record.partner_id.display_name,
                              })

    @api.model
    def check_prices(self, product, vendor_price, list_price, regular_price):
        prices_control = self.env['ir.config_parameter'].sudo().get_param('mataa_order_management.prices_control')
        if bool(prices_control) and vendor_price and vendor_price > min(list_price, regular_price):
            raise ValidationError(_(
                'Invalid purchase price detected!\n'
                'Purchase price (%(purchase_price).2f) cannot be greater than:\n'
                '- Product\'s sale price (%(sale_price).2f)\n'
                '- Product\'s regular price (%(regular_price).2f)\n'
                'Product: %(product)s\n'
            ) % {
                                      'purchase_price': vendor_price,
                                      'sale_price': list_price,
                                      'regular_price': regular_price,
                                      'product': product.display_name,
                                  })

    @api.model
    def get_file(self):
        from openpyxl import load_workbook, Workbook

        file = 'C:\Program Files\Odoo-17.0\server\odoo\mataa\mataa_product_management\models\Duplicate vendors.xlsx'
        wb = load_workbook(file)

        ws = wb.active
        col_index = 1
        ws.cell(row=1, column=col_index).value = 'ID'
        col_index += 1
        ws.cell(row=1, column=col_index).value = 'Vendor'
        col_index += 1
        ws.cell(row=1, column=col_index).value = 'Product internal Reference'
        col_index += 1
        ws.cell(row=1, column=col_index).value = 'Product name'
        col_index += 1
        ws.cell(row=1, column=col_index).value = 'Qty'
        col_index += 1
        ws.cell(row=1, column=col_index).value = 'Price'
        col_index += 1
        supplier_info_obj = self.env['product.supplierinfo']
        supplier_info_ids = supplier_info_obj.search([('product_id', '!=', False)], order="id")
        vendor_ids = set(supplier_info_ids.mapped('partner_id'))
        product_ids = set(supplier_info_ids.mapped('product_id'))
        row_index = 2
        for product in list(product_ids):
            for vendor in list(vendor_ids):
                supplier_line = supplier_info_obj.search(
                    [('product_id', '=', product.id), ('partner_id', '=', vendor.id)])
                if len(supplier_line) > 1:
                    for line in supplier_line:
                        col_index = 1
                        ws.cell(row=row_index, column=col_index).value = line.id
                        col_index += 1
                        ws.cell(row=row_index, column=col_index).value = line.partner_id.name
                        col_index += 1
                        ws.cell(row=row_index, column=col_index).value = line.product_id.default_code
                        col_index += 1
                        ws.cell(row=row_index, column=col_index).value = line.product_id.name
                        col_index += 1
                        ws.cell(row=row_index, column=col_index).value = line.min_qty
                        col_index += 1
                        ws.cell(row=row_index, column=col_index).value = line.price
                        col_index += 1
                        row_index += 1
        wb.save(file)
