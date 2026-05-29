# -*- coding: utf-8 -*-
import base64
import io
import logging

_logger = logging.getLogger(__name__)

from ..utility.import_variants_file_parser import ImportVariantsFileParser
from ..services.product_service import ProductService
from ..services.variant_service import VariantService
from ..services.vendor_service import VendorService
from ..services.category_service import CategoryService
from ..services.attribute_service import AttributeService  # لإنشاء/إسناد خصائص المتغير (مسموح)

def _norm(v):
    return (v or "").strip()

def _csv_list(v):
    if v is None:
        return []
    s = str(v).replace(";", ",").replace("|", ",")
    return [p.strip() for p in s.split(",") if p and p.strip()]


class ImportOrchestratorService:
    """
    ينسّق عملية الاستيراد:
    - يقرأ الملف عبر الـ Utility
    - يتحقق من المرجعيات الخمس (بدون أي إنشاء)
    - يمرّر القيم المعتمدة إلى الخدمات
    - يبني ملف سجل XLSX واحد لكل العملية
    """

    def __init__(self, env):
        self.env = env

    # ---------- Public API ----------
    def process_data(self, *, file_name, file_data, behaviour):
        df = self._parse_to_dataframe(file_name=file_name, file_data=file_data)

        statuses = []
        success_count = 0
        skipped_count = 0

        for idx, row in df.iterrows():
            errors = []
            try:
                with self.env.cr.savepoint():
                    refs = self._validate_refs(row=row, errors=errors)
                    if errors:
                        skipped_count += 1
                        statuses.append("\n".join(errors))
                        continue

                    self._apply_row(row=row, refs=refs, behaviour=behaviour, errors=errors)

                    if errors:
                        skipped_count += 1
                        statuses.append("\n".join(errors))
                    else:
                        success_count += 1
                        statuses.append("SUCCESS - No Errors Found")
            except Exception as e:
                _logger.exception("Unexpected error while processing row %s", idx)
                skipped_count += 1
                statuses.append(f"Unexpected Error: {str(e)}")

        has_errors = any(s != "SUCCESS - No Errors Found" for s in statuses)
        log_file_name, log_file_b64 = self._build_log_xlsx(df=df, statuses=statuses, original_name=file_name)

        return {
            "success_count": success_count,
            "skipped_count": skipped_count,
            "has_errors": has_errors,
            "log_file_name": log_file_name,
            "log_file_data": log_file_b64,
        }

    # ---------- Parse & Log ----------
    def _parse_to_dataframe(self, *, file_name, file_data):
        """
        ندعم واجهات مختلفة للـ parser (static/class or instance) لتوافق نسختكم.
        """
        # أولاً: دوال ستاتيكية/كلاسمثود مثل parse_file(...)
        for m in ("parse_file", "to_dataframe", "parse_to_dataframe", "parse"):
            if hasattr(ImportVariantsFileParser, m):
                try:
                    return getattr(ImportVariantsFileParser, m)(file_name=file_name, file_data=file_data)
                except TypeError:
                    pass  # قد تكون واجهة مختلفة

        # ثانيًا: واجهة كائن
        parser = ImportVariantsFileParser(self.env)
        for m in ("to_dataframe", "parse_to_dataframe", "parse"):
            if hasattr(parser, m):
                return getattr(parser, m)(file_name=file_name, file_data=file_data)

        raise ValueError("ImportVariantsFileParser could not parse the given file into a DataFrame.")

    def _build_log_xlsx(self, *, df, statuses, original_name):
        """
        يبني ملف XLSX واحد: كل الصفوف + Row # + Import Status مع Freeze و Wrap.
        """
        import pandas as pd
        from pandas import ExcelWriter
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        df_out = df.copy()
        df_out.insert(0, "Row #", range(1, len(df_out) + 1))
        df_out["Import Status"] = statuses

        buf = io.BytesIO()
        with ExcelWriter(buf, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="Import Log")
            ws = writer.book["Import Log"]
            # Freeze header
            ws.freeze_panes = "A2"
            # Wrap for Import Status column
            status_col_idx = list(df_out.columns).index("Import Status") + 1
            for row in ws.iter_rows(min_row=2, min_col=status_col_idx, max_col=status_col_idx, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Simple column width
            for ci in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 20

        log_name = self._derive_log_name(original_name)
        log_b64 = base64.b64encode(buf.getvalue()).decode()
        return log_name, log_b64

    def _derive_log_name(self, original):
        base = original or "import"
        return f"{base[:-5] if base.lower().endswith('.xlsx') else base}_log.xlsx"

    # ---------- Validation (NO auto-create for 5 refs) ----------
    def _validate_refs(self, *, row, errors):
        """
        يمنع الإنشاء التلقائي للخمسة:
        website categories / functional category / tags / brand / vendor
        """
        refs = {"brand": None, "func_cat": None, "web_cats": [], "tags": [], "vendor": None}

        # Functional Category (product.category) — name, =ilike
        func_name = _norm(row.get("template_functional_category"))
        if func_name:
            c = self.env["product.category"].search([("name", "=ilike", func_name)], limit=1)
            if not c:
                errors.append(f"Functional category not found: {func_name}")
            else:
                refs["func_cat"] = c

        # Website Categories (product.public.category) — CSV, all must exist
        web_csv = _norm(row.get("template_web_categories"))
        if web_csv:
            found = []
            for item in _csv_list(web_csv):
                leaf = CategoryService.get_category_by_complete_path(self.env, item)
                if leaf:
                    found.append(leaf)
            refs["web_cats"] = found

        # Tags (product.tag) — CSV, case-insensitive
        tag_csv = _norm(row.get("template_tags"))
        if tag_csv:
            names = _csv_list(tag_csv)
            if names:
                existing = self.env["product.tag"].search([("name", "in", names)])
                map_l = {t.name.lower().strip(): t for t in existing}
                missing = [n for n in names if n.lower().strip() not in map_l]
                if missing:
                    errors.append("Tags not found: " + ", ".join(missing))
                else:
                    refs["tags"] = [map_l[n.lower().strip()] for n in names]

        # Brand (product.brand)
        brand_name = _norm(row.get("template_brand"))
        if brand_name:
            b = self.env["product.brand"].search([("name", "=ilike", brand_name)], limit=1)
            if not b:
                errors.append(f"Brand not found: {brand_name}")
            else:
                refs["brand"] = b

        # Vendor (res.partner) — NO supplier_rank condition per your latest note
        vendor_name = _norm(row.get("variant_vendor_name"))
        vendor_price = row.get("variant_vendor_price")
        if vendor_name and (vendor_price is not None):
            v = self.env["res.partner"].search([("name", "=ilike", vendor_name)], limit=1)
            if not v:
                errors.append(f"Vendor not found: {vendor_name}")
            else:
                refs["vendor"] = v

        return refs

    # ---------- Apply (use existing Services only) ----------
    def _apply_row(self, *, row, refs, behaviour, errors):
        env = self.env

        # Template identity (required)
        tmpl_code = _norm(row.get("template_internal_ref") or row.get("default_code"))
        if not tmpl_code:
            errors.append("Missing template_internal_ref (default_code).")
            return

        tmpl_name = _norm(row.get("template_name"))
        tmpl_regular = row.get("template_regular_price")
        tmpl_sales = row.get("template_sales_price")
        tmpl_desc = row.get("template_Description") or row.get("template_description")
        tmpl_note = row.get("template_internal_note")

        # Images (reuse ProductService helpers)
        product_urls = []
        for img_url in (self._parser_images(row) or []):
            if not img_url:
                continue
            url = img_url.strip() if isinstance(img_url, str) else str(img_url).strip()
            if not url:
                continue
            stream = ProductService.get_stream(url)
            if not stream:
                continue
            vals = ProductService.get_image_url_vals(url, stream)
            vals.update({"sequence": len(product_urls) + 1})
            product_urls.append((0, 0, vals))

        # Template create/update via ProductService
        tmpl = ProductService.get_product(env=env, default_code=tmpl_code)
        vals_common = {
            "product_name": tmpl_name or (tmpl.name if tmpl else tmpl_code),
            "default_code": tmpl_code,
            "regular_price": tmpl_regular,
            "sales_price": tmpl_sales,
            "category_id": refs["func_cat"].id if refs["func_cat"] else None,
            "description_sale": tmpl_desc,
            "internal_note": tmpl_note,
            "image_url_ids": product_urls,
            "brand_id": refs["brand"].id if refs["brand"] else None,
        }
        if tmpl:
            tmpl = ProductService.update_product(env=env, **vals_common)
        else:
            if behaviour == "only_update":
                errors.append(
                    f"Product Template isn't found for Internal Ref {tmpl_code}, "
                    f"please change behaviour to create/update to create a new template."
                )
                return
            tmpl = ProductService.create_product(env=env, **vals_common)

        # Assign website categories (existing only)
        if refs["web_cats"]:
            tmpl.sudo().write({"public_categ_ids": [(6, 0, [c.id for c in refs["web_cats"]])]})

        # Assign template tags (existing only)
        if refs["tags"]:
            tmpl.sudo().write({"product_tag_ids": [(6, 0, [t.id for t in refs["tags"]])]})

        # Variant (no creation except attribute values allowed in create_update)
        var_code = _norm(row.get("variant_internal_ref"))
        var_name = _norm(row.get("variant_name"))
        var_regular = row.get("variant_regular_price")
        var_barcodes = _csv_list(row.get("variant_barcodes")) if row.get("variant_barcodes") else None

        attribute_value_ids = set()
        if behaviour == "create_update":
            for attr_name, value_name in self._parser_attribute_pairs(row):
                a = AttributeService.get_or_create_attribute(env, _norm(attr_name))
                v = AttributeService.get_or_create_attribute_value(env, a, _norm(value_name))
                ProductService.assign_product_attribute_values(env, tmpl, a, v)
                attribute_value_ids.add(v.id)

        if var_code:
            if behaviour == "create_update":
                variant = VariantService.update_variant_with_attributes(
                    env=env,
                    product_template=tmpl,
                    variant_name=var_name,
                    internal_ref=var_code,
                    regular_price=var_regular,
                    barcodes=var_barcodes,
                    attribute_values=attribute_value_ids or None,
                )
            else:
                variant = VariantService.update_variant(
                    env=env,
                    product_template=tmpl,
                    variant_name=var_name,
                    internal_ref=var_code,
                    regular_price=var_regular,
                    barcodes=var_barcodes,
                )

            # Variant tags (existing only)
            var_tag_csv = _norm(row.get("variant_tags"))
            if var_tag_csv:
                names = _csv_list(var_tag_csv)
                existing = env["product.tag"].search([("name", "in", names)])
                map_l = {t.name.lower().strip(): t for t in existing}
                missing = [n for n in names if n.lower().strip() not in map_l]
                if missing:
                    errors.append("Missing Variant Tags: " + ", ".join(missing))
                    return
                variant.sudo().write({"product_tag_ids": [(6, 0, [map_l[n.lower().strip()].id for n in names])]})

            # Vendor (existing only) + price checks
            vend_name = _norm(row.get("variant_vendor_name"))
            vend_price = row.get("variant_vendor_price")
            vend_prod_name = row.get("variant_vendor_product_name")
            vend_code = row.get("variant_vendor_product_code")
            vend_qty = row.get("variant_vendor_quantity")

            if vend_name and (vend_price is not None):
                vendor = refs["vendor"]
                if not vendor:
                    errors.append(f"Vendor not found: {vend_name}")
                    return

                VendorService.link_vendor_to_variant(
                    env,
                    vendor=vendor,
                    product_variant=variant,
                    vendor_product_name=vend_prod_name,
                    vendor_code=vend_code,
                    vendor_price=vend_price,
                    vendor_quantity=vend_qty,
                )
                env["product.supplierinfo"].check_prices(
                    variant, vend_price, tmpl_sales, var_regular
                )

    # ---------- Utility parsers (defer to ImportVariantsFileParser) ----------
    def _parser_images(self, row):
        for m in ("parse_product_images", "get_product_images", "product_images"):
            if hasattr(ImportVariantsFileParser, m):
                try:
                    result = getattr(ImportVariantsFileParser, m)(row)
                    return result
                except Exception as e:
                    _logger.warning(f"Error parsing images using method {m}: {str(e)}")
                    continue
        return None

    def _parser_attribute_pairs(self, row):
        for m in ("parse_attribute_values", "get_attribute_values", "attribute_pairs"):
            if hasattr(ImportVariantsFileParser, m):
                try:
                    return getattr(ImportVariantsFileParser, m)(row) or []
                except Exception:
                    pass
        return []
